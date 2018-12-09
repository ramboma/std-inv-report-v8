import os
import pandas as pd
import openpyxl as xl
from openpyxl.styles import numbers as numStyle
import data_analysis.config as CONFIG


class AnalysisResultWriter(object):
    def __init__(self, folder):
        self._folder = folder

    @property
    def folder(self):
        if self._folder[-1] != os.sep:
            self._folder = self._folder + os.sep
        else:
            self._folder = self._folder
        return self._folder

    def write_new_sheet(self, book_name, sheet_name, df):
        try:
            file = os.path.join(self.folder, book_name)
            writer = pd.ExcelWriter(file)
            if isinstance(df.columns, pd.MultiIndex):
                index = True
            else:
                index = False
            if not os.path.exists(file):
                df.to_excel(writer, sheet_name, index=index)
            else:
                book = xl.load_workbook(writer.path)
                writer.book = book
                df.to_excel(excel_writer=writer, sheet_name=sheet_name, index=index)
            writer.save()
        except Exception as e:
            raise ("文件写入时，发生异常，异常：{}".format(e.__str__()))
        finally:
            writer.close()

    def write_new_book(self, book_name, df_dict):
        if df_dict:
            for sheet_name in df_dict:
                self.write_new_sheet(book_name, sheet_name, df_dict[sheet_name])

            self.formate_percent(book_name)

    def formate_percent(self, book_name):
        try:
            file = os.path.join(self.folder, book_name)

            wbook = xl.load_workbook(file)
            sheet_names = wbook.sheetnames
            for sheet_name in sheet_names:
                sheet = wbook[sheet_name]
                merg_cells = sheet.merged_cells
                merg_cells_range=sheet.merged_cells.ranges
                max_row = sheet.max_row
                max_col = sheet.max_column

                for i in range(1, max_col + 1):
                    colTag = xl.utils.get_column_letter(i)
                    sheet.column_dimensions[colTag].width = 10
                    if merg_cells_range:
                        cell_v=None
                        for merg in merg_cells_range:
                            if i >= merg.min_col and i<=merg.max_col:
                                cell_v=sheet.cell(row=1, column=merg.min_col).value
                                break
                        if cell_v is None:
                            cell_v = sheet.cell(row=1, column=i).value
                    else:
                        cell_v = sheet.cell(row=1, column=i).value
                    if str(cell_v).find("人数") >= 0:
                        pass
                    elif str(cell_v).find("均值") >= 0:
                        sheet.column_dimensions[colTag].number_format = numStyle.FORMAT_NUMBER_00
                        for j in range(1 + 1, max_row + 1):
                            sheet.cell(row=j, column=i).number_format = numStyle.FORMAT_NUMBER_00
                    else:
                        sheet.column_dimensions[colTag].number_format = numStyle.FORMAT_PERCENTAGE_00
                        for j in range(1 + 1, max_row + 1):
                            sheet.cell(row=j, column=i).number_format = numStyle.FORMAT_PERCENTAGE_00

                wbook.save(file)
        except:
            pass
        finally:
            wbook.close()
