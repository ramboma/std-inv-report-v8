#!/usr/bin/env python3
# -*- coding: utf-8 -*-

'read_excel_util.py'

__author__ = 'kuoren'

import pandas as pd
import openpyxl as xl
from openpyxl.styles import numbers as numStyle
import os
import data_analysis.config as CONFIG


def read_excel(filePath):
    '''
    Read excel
    :param filePath:
    :return:
    '''
    print("loading file...")
    xls = pd.ExcelFile(filePath)
    df = xls.parse()
    return df


def writeExcel(dataFrame, filePath, sheetName):
    if dataFrame is None or dataFrame.empty:
        return
    writer = pd.ExcelWriter(filePath)
    if os.path.exists(filePath) != True:
        dataFrame.to_excel(writer, sheetName, index=None)
    else:
        book = xl.load_workbook(writer.path)
        writer.book = book
        dataFrame.to_excel(excel_writer=writer, sheet_name=sheetName, index=None)
    writer.save()
    writer.close()

    cols=[col for col in dataFrame.columns]
    percent_cols=percent_columns(cols)
    formate_percent(filePath,sheetName,percent_cols)

def writeExcelWithIndex(dataFrame, filePath, sheetName):
    if dataFrame is None or dataFrame.empty:
        return

    writer = pd.ExcelWriter(filePath)
    if os.path.exists(filePath) != True:
        dataFrame.to_excel(writer, sheetName, index=True)
    else:
        book = xl.load_workbook(writer.path)
        writer.book = book
        dataFrame.to_excel(excel_writer=writer, sheet_name=sheetName, index=True)
    writer.save()
    writer.close()

    percent_cols = [CONFIG.MEASURE_NAME_HELP,CONFIG.MEASURE_NAME_SATISFY,CONFIG.MEASURE_NAME_MEET]
    formate_percent(filePath, sheetName, percent_cols,2)


def formate_percent(file_path, sheet_name, percent_cols,head=1):
    wbook = xl.load_workbook(file_path)
    sheet = wbook[sheet_name]
    max_row = sheet.max_row
    max_col = sheet.max_column

    for i in range(1, max_col + 1):
        colTag = xl.utils.get_column_letter(i)
        sheet.column_dimensions[colTag].width = 10

        if sheet.cell(row=head, column=i).value in percent_cols:
            sheet.column_dimensions[colTag].number_format = numStyle.FORMAT_PERCENTAGE_00
            for j in range(head+1, max_row + 1):
                sheet.cell(row=j, column=i).number_format = numStyle.FORMAT_PERCENTAGE_00
    wbook.save(file_path)
    wbook.close()


def percent_columns(columns):
    elimite_cols = [CONFIG.MEAN_COLUMN[2], CONFIG.MEAN_COLUMN[-1], CONFIG.MEAN_COLUMN[1],
                    CONFIG.MEAN_COLUMN[0], CONFIG.ABILITY_COLUMN,CONFIG.GROUP_COLUMN[2],
                    CONFIG.GROUP_COLUMN[0], CONFIG.GROUP_COLUMN[1],  CONFIG.TOTAL_COLUMN]
    for column in elimite_cols:
        if column in columns:
            columns.remove(column)
    subs = columns[0:]
    for column in subs:
        if column.find(CONFIG.MEAN_COLUMN[2]) >= 0 or column.find(CONFIG.MEAN_COLUMN[-1]) >= 0:
            columns.remove(column)

    return columns


if __name__ == "__main__":
    read_excel(
        "../test-data/san-ming/cleaned/AnswerList1540806254513_cleaned_本科毕业生_public_analysis_20181102235123.xlsx")
