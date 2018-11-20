#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""utils.py"""

__author__ = 'Gary.Z'

import openpyxl as xl
import re

from itertools import product
from data_cleansing.config import *

logger = get_logger(__name__)


def generate_excel_column_indexes(seed=list('ABCDEFGHIJKLMNOPQRSTUVWXYZ'), iter_cnt=1):
    col_lst = []
    for index in range(1, iter_cnt + 1):
        lst = list(product(seed, repeat=index))  # 得到排列序元组序列
        lst = map(lambda elem: ''.join(elem), lst)  # 将排列元组序列转成字符串序列
        lst = list(set(lst))  # 消除重复元素
        lst = sorted(lst)  # 按字母ASCII的顺序进行排列
        col_lst += lst
    return col_lst


def add_item_to_list_dict(dict, key, value):
    if key not in dict:
        dict[key] = []
    if value not in dict[key]:
        dict[key].append(value)


# def validate_data_dimensions(work_sheet, expect_cols=231, expect_rows=3):
#     """data dimension checking: row >=3 and col >= 231 """
#     logger.info('validating data dimensions, cols: {}, rows: {}'.format(work_sheet.max_column, work_sheet.max_row))
#     if work_sheet.max_column < expect_cols:
#         raise Exception("column count must >= {}".format(expect_cols))
#     if work_sheet.max_row < 3:
#         raise Exception("row count must >= {}".format(expect_rows))


# def validate_question_id_header(values):
#     pattern = r'(?P<prefix>([A-Z0-9]+)(-[0-9]+)*)(-[A-Z]+)?'
#     for value in values:
#         if value is None or value == '':
#             continue
#         matches = re.match(pattern, value)
#         if matches is None:
#             raise Exception('invalid question id: {}, consider missing question id header'.format(value))


def extract_question_id_prefix(title):
    matches = re.match(r'(?P<prefix>([A-Z0-9]+)(-[0-9]+)*)(-[A-Z]+)?', title)
    if matches is None:
        return title
    else:
        return matches.group('prefix')


def reset_column_names(header_cells, excel_column_indexes):

    # Set question-answers column headers
    answer_column = False
    flag1 = False
    flag2 = False
    for i in range(0, header_cells.__len__() - 1):
        header_name = header_cells[i]

        if not answer_column:
            if header_name is None or header_name == '':
                header_name = '_' + str(i + 1)
                header_cells[i] = header_name
                continue

            if not answer_column and header_name == 'A1':
                answer_column = True
                if i < BASE_INFO_COLS_MIN:
                    raise Exception("base info columns must >= 23, current: {}".format(i))

        next_header_name = header_cells[i + 1]

        if i >= header_cells.__len__() - 2:
            a = 1
            pass

        if header_name is not None and next_header_name is None:
            flag2 = True
            prefix = header_name
            option = 0

        if header_name is None and next_header_name is not None:
            flag1 = True

        if flag2:
            new_header_name = '{}-{}'.format(prefix, excel_column_indexes[option])
            option += 1
            header_cells[i] = new_header_name

        if flag1:
            flag1 = False
            flag2 = False
            prefix = ''
            option = 1

    if next_header_name is None:
        new_header_name = '{}-{}'.format(prefix, excel_column_indexes[option])
        header_cells[-1] = new_header_name


def build_question_to_column_mapping_v2(column_names, mapping):
    mapping.clear()
    idx = 0
    for column_name in column_names:
        prefix_name = extract_question_id_prefix(column_name)

        add_item_to_list_dict(mapping, prefix_name, idx)
        add_item_to_list_dict(mapping, column_name, idx)

        idx += 1

    return mapping


def build_question_to_column_mapping(work_sheet, excel_column_indexes):
    mapping = {}
    header_cells = work_sheet[1]
    for header_cell in header_cells:
        header_name = header_cell.value
        header_prefix = extract_question_id_prefix(header_name)
        column_name = excel_column_indexes[header_cell.col_idx - 1]

        add_item_to_list_dict(mapping, header_prefix, column_name)
        add_item_to_list_dict(mapping, header_name, column_name)
    return mapping


def add_tracing_comment(cell, rule_no, func, addition=None):
    if addition is None:
        text = 'rule {}\norigin val: {}\nfunc: {}'.format(rule_no, cell.value, func)
    else:
        text = 'rule {}\norigin val: {}\nfunc: {}\n{}'.format(rule_no, cell.value, func, addition)
    cell.comment = xl.comments.Comment(text, None, 150, 300)


def remove_rows_by_index_list(work_sheet, index_list, rule, func, trace_mode=False):
    for i in range(0, index_list.__len__())[::-1]:
        if trace_mode:
            for cell in work_sheet[index_list[i]]:
                add_tracing_comment(cell, rule, func)
        else:
            work_sheet.delete_rows(index_list[i])
    logger.info('>> {} rows removed'.format(index_list.__len__()))
    # logger.debug('>> {}'.format(index_list))


def rinse_values_by_column_rowindex(work_sheet, col, index_list, rule, func, debug=False, trace_mode=False):
    _debug_info_ = []
    counter = 0
    for i in index_list:
        coordinate = '{}{}'.format(col, i)
        if work_sheet[coordinate].value is not None:
            if debug:
                _debug_info_.append(work_sheet[coordinate].value)
            if trace_mode:
                add_tracing_comment(work_sheet[coordinate], rule, func)
            else:
                work_sheet[coordinate].value = None
            counter += 1
    logger.info('>> {} cells rinsed'.format(counter))
    if debug:
        logger.debug('>> {}'.format(_debug_info_))


def rinse_values_by_column(work_sheet, col, rule, func, debug=False, trace_mode=False):
    _debug_info_ = []
    i = 0
    for cell in work_sheet[col]:
        if cell.value is not None:
            if debug:
                _debug_info_.append(cell.value)
            if trace_mode:
                add_tracing_comment(cell, rule, func)
            else:
                cell.value = None
        i += 1
    logger.info('>> {} cells rinsed'.format(i))
    if debug:
        logger.debug('>> {}'.format(_debug_info_))


def query_row_indexes_by_column_filter(work_sheet, xl_col, cb_filter):
    idx_list = []
    for cell in work_sheet[xl_col]:
        if cell.row <= HEADER_ROW_INDEX:
            continue
        if cb_filter(cell.value):
            idx_list.append(cell.row)
    return idx_list


def test():
    pass


if __name__ == '__main__':
    test()
