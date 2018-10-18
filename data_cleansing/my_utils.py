#!/usr/bin/env python3
# -*- coding: utf-8 -*-

'utils.py'

__author__ = 'Gary.Z'

# import xlrd
# import xlwt
import openpyxl as xl
import numpy as np

import pandas as pd

INDEX_BASE = 1
HEADER_ROW_INDEX = 0 + INDEX_BASE
# MAJOR_COLUMN_INDEX = 13 + INDEX_BASE
# SUBMIT_TIME_COLUMN_INDEX = 21 + INDEX_BASE
A1_COLUMN_INDEX = 23 + INDEX_BASE
# A2_COLUMN_INDEX = 25 + INDEX_BASE
# G1_COLUMN_INDEX = 60 + INDEX_BASE

# MAJOR_COLUMN_EXCEL_INDEX = chr(ord('A') + MAJOR_COLUMN_INDEX - INDEX_BASE)
# SUBMIT_TIME_COLUMN_EXCEL_INDEX = chr(ord('A') + SUBMIT_TIME_COLUMN_INDEX - INDEX_BASE)
# A1_COLUMN_EXCEL_INDEX = chr(ord('A') + A1_COLUMN_INDEX - INDEX_BASE)
# A2_COLUMN_EXCEL_INDEX = chr(ord('A') + A2_COLUMN_INDEX - INDEX_BASE)
# G1_COLUMN_EXCEL_INDEX = chr(ord('A') + G1_COLUMN_INDEX - INDEX_BASE)
MAJOR_COLUMN_EXCEL_INDEX = 'N'
SUBMIT_TIME_COLUMN_EXCEL_INDEX = 'V'
A1_COLUMN_EXCEL_INDEX = 'X'
A2_COLUMN_EXCEL_INDEX = 'Z'
B6_COLUMN_EXCEL_INDEX = 'AL'
G1_COLUMN_EXCEL_INDEX_A = 'BH'
G1_COLUMN_EXCEL_INDEX_B = 'BI'
H5_COLUMN_EXCEL_INDEX_NC = 'DV'
H6_COLUMN_EXCEL_INDEX_NC = 'ED'

MAJOR_FILTER_LIST = ('测试专业')
NC_OPTION_FILTER_LIST = ('无法评价', '以上均不需要改进')
G1_OPTION_FILTER_LIST = ('国际组织', '军队')


def validate_data_dimensions(st):
    """rule 0: data dimension checking: row >=3 and col >= 231 """
    print('rule 0: validating data dimensions, cols: {}, rows: {}'.format(st.max_column, st.max_row))
    if st.max_column < 231:
        raise Exception("column count must >= 231")
    if st.max_row < 3:
        raise Exception("row count must >= 3")


def remove_unnecessary_headers(st):
    """rule 0: remove row 1~2: include question description and option description"""
    print('rule 0: removing unnecessary header rows start at {}, count 2'.format(HEADER_ROW_INDEX + 1))
    st.delete_rows(HEADER_ROW_INDEX + 1, 2)


def batch_reset_column_names(st):
    """rule 0: set first 23 column name set with _1~_23, rest set follow predefined rules, e.g. A1-A"""
    print('rule 0: batch reset column names with standard codes')
    BOUNDARY_0 = 0 + INDEX_BASE
    BOUNDARY_1 = 23 + INDEX_BASE
    BOUNDARY_2 = (st.max_column - 1) + INDEX_BASE

    # Set base info column headers
    for i in range(BOUNDARY_0, BOUNDARY_1):
        value = '_' + str(i)
        st.cell(HEADER_ROW_INDEX, i, value)
        # print('write: "' + value + '"')

    # Set question-answers column headers
    value = ''
    next_value = ''
    flag1 = False
    flag2 = False
    for i in range(BOUNDARY_1, BOUNDARY_2):  # loop from index 23 to last - 1
        value = st.cell(HEADER_ROW_INDEX, i).value
        next_value = st.cell(HEADER_ROW_INDEX, i + 1).value

        if value is not None and next_value is None:
            flag2 = True
            prefix = value
            option = 'A'

        if value is None and next_value is not None:
            flag1 = True

        if flag2:
            value_to_write = prefix + '-' + option
            option = chr(ord(option) + 1)
            st.cell(HEADER_ROW_INDEX, i, value_to_write)
            # print('write: "' + value_to_write + '"')
        # else:
        #     value_to_write = value

        if flag1:
            flag1 = False
            flag2 = False
            prefix = ''
            option = 'A'

    st.title = "cleaned"


def query_row_indexes_by_column_filter(st, xl_col, cb_filter):
    idx_list = []
    for cell in st[xl_col]:
        if cell.row <= HEADER_ROW_INDEX:
            continue
        # print("cell: {}".format(cell.value))
        if cb_filter(cell.value):
            idx_list.append(cell.row)
    # print(idx_list)
    return idx_list


def remove_rows_by_index_list(st, index_list):
    for i in range(0, index_list.__len__())[::-1]:
        st.delete_rows(index_list[i])
        # print('remove row: {}'.format(index_list[i]))
    print('>> {} rows removed'.format(index_list.__len__()))


def remove_fake_records(st):
    """rule 1: remove fake data, e.g. column 14(专业名称) with value "测试专业" """
    print('rule 1: removing rows which major in {}'.format(MAJOR_FILTER_LIST))
    # find them
    remove_list = query_row_indexes_by_column_filter(st, MAJOR_COLUMN_EXCEL_INDEX, lambda val: val in MAJOR_FILTER_LIST)
    # remove them
    remove_rows_by_index_list(st, remove_list)


def remove_unsubmitted_records(st):
    """rule 2, 3: remove un-submitted row, e.g. no submit-time exist"""
    print('rule 2, 3: removing rows which have no submit time')
    # find them
    remove_list = query_row_indexes_by_column_filter(st, SUBMIT_TIME_COLUMN_EXCEL_INDEX, lambda val: (val is None or val == ''))
    # remove them
    remove_rows_by_index_list(st, remove_list)


def remove_unqualified_records(st):
    """rule 2, 3: remove un-qualified row, e.g. no answer for question A2"""
    print('rule 2, 3: removing rows which have no A2 answers')
    # find them
    remove_list = query_row_indexes_by_column_filter(st, A2_COLUMN_EXCEL_INDEX, lambda val: (val is None or val == ''))
    # remove them
    remove_rows_by_index_list(st, remove_list)


def rinse_nonrelevance_answers(df):
    """rule 4: replace non-relevance answers(cell) with NaN against question-relevance rules"""
    # IF A2 not in (在国内工作, 自由职业) then rinse B1, B2, B3, B4, B5, B6, B7, B8, B9-1, B9-2, B10-1, B10-2, D1, D2
    # IF A2 = 自由职业 then rinse B1,B2,B3,B4, B10-1
    # IF B9-1 not in (比较不相关, 很不相关) then rinse B9-2
    # IF B10-1 = 0次 then rinse B10-2
    # IF A2 != 未就业 then rinse C1, C2
    # IF A2 != 在国内升学 then rinse E1,E2,E3,E4
    # IF E3 not in (比较不相关, 很不相关) then rinse E4
    # IF A4 != 出国/出境 then rinse F1, F2, F3, F4
    # IF F1 != 求学 then rinse F2, F3
    # IF F3 not in (比较不相关, 很不相关) then rinse F4
    # IF A4 != 自主创业 then rinse G1, G2, G3, G4, G5


def rinse_nc_option_values(st):
    """rule 5: replace values like "无法评价", "以上均不需要改进" with NaN """
    print('rule 5: rinse answers which in {} into NaN'.format(NC_OPTION_FILTER_LIST))
    i = 0
    for row in range(HEADER_ROW_INDEX + 1, st.max_row + 1):
        for col in range(A1_COLUMN_INDEX, st.max_column + 1):
            if st.cell(row, col).value in NC_OPTION_FILTER_LIST:
                cell = st.cell(row, col, None)
                # print('rinse cell: {} - {}'.format(cell.coordinate, cell.value))
                i += 1
    print('>> {} cells rinsed'.format(i))

    rinse_values_by_column_rowindex(st, H5_COLUMN_EXCEL_INDEX_NC, range(HEADER_ROW_INDEX + 1, st.max_row + 1))
    rinse_values_by_column_rowindex(st, H6_COLUMN_EXCEL_INDEX_NC, range(HEADER_ROW_INDEX + 1, st.max_row + 1))


def rinse_values_by_column_rowindex(st, col, index_list):
    for i in index_list:
        coordinate = '{}{}'.format(col, i)
        if st[coordinate] is not None:
            # print('rinse cell: {} - {}'.format(coordinate, st[coordinate].value))
            st[coordinate] = None
            i += 1
    print('>> {} cells rinsed'.format(index_list.__len__()))


def rinse_invalid_answers(st):
    """rule 6: replace invalid answers(cell) with NaN"""
    print('rule 6: rinse G1 answers which in {}'.format(G1_OPTION_FILTER_LIST))
    # find them
    rinse_list = query_row_indexes_by_column_filter(st, G1_COLUMN_EXCEL_INDEX_A,
                                                    lambda val: val in G1_OPTION_FILTER_LIST)
    # remove them
    rinse_values_by_column_rowindex(st, G1_COLUMN_EXCEL_INDEX_A, rinse_list)
    rinse_values_by_column_rowindex(st, G1_COLUMN_EXCEL_INDEX_B, rinse_list)


def rinse_unusual_salary_values(st):
    """rule 7: remove < 1000, top 0.3%, ABS(diff of MEAN) > 4 * STDEV """
    print('rule 7: remove < 1000, top 0.3%, ABS(diff of MEAN) > 4 * STDEV ')
    print('>> rinsing salary < 1000')
    rinse_list = query_row_indexes_by_column_filter(st, B6_COLUMN_EXCEL_INDEX, filter_low_salary)
    rinse_values_by_column_rowindex(st, B6_COLUMN_EXCEL_INDEX, rinse_list)

    print('>> rinsing top N salary')
    sort_range = '{}{}:{}{}'.format(B6_COLUMN_EXCEL_INDEX, 2, B6_COLUMN_EXCEL_INDEX, st.max_row);
    # st.auto_filter.add_sort_condition(sort_range)
    salary_list = {}
    for row in st[sort_range]:
        if row[0].value is not None and row[0].value != '':
            salary_list[row[0].coordinate] = int(row[0].value)

    sorted_salary_list = sorted(salary_list.items(), key=lambda kv: kv[1], reverse=True)
    top_n = round(sorted_salary_list.__len__() * 0.003)
    i = 0
    for item in sorted_salary_list:
        coordinate = item[0]
        # print('rinse cell: {} - {}'.format(coordinate, st[coordinate].value))
        st[coordinate] = None
        salary_list.pop(coordinate)
        i += 1
        if i >= top_n:
            break
    print('>> {} cells rinsed from {}'.format(top_n, sorted_salary_list.__len__()))

    print('>> rinsing ABS(salary - MEAN) > 4 * STDEV')
    np_salary_list = np.array(list(salary_list.values()), dtype=int)
    print('>> {} values in total'.format(np_salary_list.size))
    salary_mean = np_salary_list.mean()
    print('>> MEAN = {}'.format(salary_mean))
    salary_stdev = np_salary_list.std()
    print('>> STDEV = {}'.format(salary_stdev))
    salary_stdev_4 = salary_stdev * 4
    print('>> * 4 = {}'.format(salary_stdev_4))
    i = 0
    for coordinate in salary_list:
        salary = int(st[coordinate].value)
        if abs(salary - salary_mean) > salary_stdev_4:
            # print('rinse cell: {} - {}'.format(coordinate, st[coordinate].value))
            st[coordinate] = None
            i += 1
    print('>> {} cells rinsed'.format(i))


def filter_low_salary(val):
    if val is None or val == '':
        return False
    result = False;
    try:
        s = int(val)
        result = s < 1000
    except ValueError as e:
        print('>> failed to process {} - {}'.format(val, e))
    finally:
        pass
    return result


def test():
    wb = xl.load_workbook('../test-data/san-ming/raw/answer20181016_1740112347.xlsx')
    st = wb.worksheets[0]
    validate_data_dimensions(st)
    batch_reset_column_names(st)
    remove_unnecessary_headers(st)
    # Rule 1
    remove_fake_records(st)
    # Rule 2, 3
    remove_unsubmitted_records(st)
    # Rule 2, 3
    remove_unqualified_records(st)
    # Rule 5
    rinse_nc_option_values(st)
    # Rule 6
    rinse_invalid_answers(st)
    # Rule 7
    rinse_unusual_salary_values(st)
    wb.save('../result/cleaned/answer20181016_1740112347_cleaned.xlsx')

if __name__=='__main__':
    test()