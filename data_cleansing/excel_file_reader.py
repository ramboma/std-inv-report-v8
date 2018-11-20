#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""cleanser_runner.py"""

__author__ = 'Gary.Z'

import time
import xlrd
import openpyxl as xl

from data_cleansing.config import *

logger = get_logger(__name__)


class ExcelFileProcessor:
    def __init__(self, file):
        self._file = file
        self._xl_reader = None
        self._logger = get_logger('{}${}'.format(self.__class__.__name__, id(self)))

    def try_load(self):
        self._xl_reader = ExcelFileOpenXLReader(self._file)
        self._logger.debug('Try load excel file using ExcelFileOpenXLReader')
        if self._xl_reader.try_load():
            self._logger.debug('done')
            return True

        self._xl_reader = ExcelFileXlrdReader(self._file)
        self._logger.debug('Try load excel file using ExcelFileXlrdReader')
        if self._xl_reader.try_load():
            self._logger.debug('done')
            return True

        raise Exception('failed to load excel file, all read methods failed')

    def traverse_rows(self, process_row, args):
        return self._xl_reader.traverse_rows(process_row, args)

    def get_max_cols(self):
        return self._xl_reader.get_max_cols()

    def get_max_rows(self):
        return self._xl_reader.get_max_rows()

    def close(self):
        self._xl_reader.close()


class ExcelFileReader:
    def __init__(self, file):
        self._file = file
        self._workbook = None
        self._max_cols = 0
        self._max_rows = 0
        # self._logger = get_logger('{}${}'.format(self.__class__.__name__, id(self)))

    def _try_open_workbook(self, max_retry=20, retry_interval=3):
        wb = None
        success = False
        retry_count = max_retry
        while (not success) and retry_count > 0:
            try:
                wb = self._open_workbook()
                success = True
            except Exception as e:
                logger.debug(e)
                logger.info("waiting for workbook lock release")
                time.sleep(retry_interval)
            finally:
                retry_count -= 1
        if not success:
            raise Exception('workbook is locked by another process and exceeded waiting time limit: {} secs'.format(
                retry_interval * max_retry))

        self._workbook = wb

    def _open_workbook(self):
        raise Exception('not implement')

    def _get_worksheet(self):
        raise Exception('not implement')

    def _validate_worksheet(self):
        raise Exception('not implement')

    def try_load(self):
        self._try_open_workbook()
        return self._validate_worksheet()

    def traverse_rows(self, process_row, args):
        raise Exception('not implement')

    def get_max_cols(self):
        return self._max_cols

    def get_max_rows(self):
        return self._max_rows

    def _copy_row_as_values(self, row):
        raise Exception('not implement')

    def close(self):
        raise Exception('not implement')


class ExcelFileOpenXLReader(ExcelFileReader):
    def __init__(self, file):
        super().__init__(file)

    def _open_workbook(self):
        return xl.load_workbook(self._file, True)

    def _get_worksheet(self):
        return self._workbook.worksheets[0]

    def _validate_worksheet(self):
        ws = self._get_worksheet()
        # ws.calculate_dimension(True)
        self._max_cols = ws.max_column
        self._max_rows = ws.max_row

        if ws.max_column <= 1 and ws.max_row <= 1:
            return False
        else:
            return True

    def traverse_rows(self, process_row, args):
        ws = self._get_worksheet()
        row_num = 0
        for row in ws.rows:
            row_num += 1
            process_row(row_num, row, self._copy_row_as_values(row), args[0], args[1], args[2])
        #     if row_num % 1000 == 0:
        #         report_progress(row_num)
        # report_progress(row_num)

        return row_num

    def _copy_row_as_values(self, row):
        value_list = []
        for i in range(0, self._max_cols):
            if i < row.__len__():
                value = row[i].value
                if value == '':
                    value = None
            else:
                value = None
            value_list.append(value)
        return value_list

    def close(self):
        self._workbook.close()


class ExcelFileXlrdReader(ExcelFileReader):
    def __init__(self, file):
        super().__init__(file)

    def _open_workbook(self):
        return xlrd.open_workbook(self._file, on_demand=True)

    def _get_worksheet(self):
        return self._workbook.sheet_by_index(0)

    def _validate_worksheet(self):
        ws = self._get_worksheet()

        self._max_cols = ws.ncols
        self._max_rows = ws.nrows

        if ws.ncols <= 1 and ws.nrows <= 1:
            return False
        else:
            return True

    def traverse_rows(self, process_row, args):
        ws = self._get_worksheet()
        for row_num in range(ws.nrows):
            process_row(row_num + 1, ws.row(row_num), self._copy_row_as_values(ws.row(row_num)), args[0], args[1], args[2])
        #     if row_num % 1000 == 0:
        #         report_progress(row_num)
        # report_progress(row_num)

        return ws.nrows

    def _copy_row_as_values(self, row):
        value_list = []
        for i in range(0, self._max_cols):
            if i < row.__len__():
                value = row[i].value
                if value == '':
                    value = None
            else:
                value = None
            value_list.append(value)
        return value_list

    def close(self):
        self._workbook.release_resources()
